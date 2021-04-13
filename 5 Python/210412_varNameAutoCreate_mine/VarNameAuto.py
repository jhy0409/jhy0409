import openpyxl
import pyautogui
import pyperclip
import time

wb = openpyxl.Workbook()
sheet = wb.active
makebyonsu = 'makeButton_'
makeNum = 150

sheet['A1'] = '변수명'
sheet['C1'] = '생성개수'
sheet['B1'] = makebyonsu
sheet['D1'] = makeNum

# makeSu = sheet['D1']
# num = (int)makeSu+1
i = 1;
for item in range(1,makeNum+1):
    sheet.cell(row=i, column=6).value = i
    i = i+1

idex = 1;
for item in range(1,makeNum+1):
    strI = str(idex)
    sheet.cell(row=idex, column=8).value = makebyonsu + strI
    idex = idex+1

idex = 1;
allVarName = '';
for item in range(idex,makeNum+1):
    allVarName = allVarName+sheet.cell(row=idex, column=8).value + ', '
    idex = idex+1

sheet['J1'] = allVarName

wb.save('test2.xlsx')


# sheet.cell(row=3, column=3).value = '3, 3'
# sheet.append([1, 2, 3, 4, 5])
# sheet['F1'] = '변수명'